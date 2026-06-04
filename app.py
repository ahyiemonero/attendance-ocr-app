import os
import re
import uuid
import sqlite3
import threading
from datetime import datetime

from flask import Flask, render_template, request, redirect, url_for, send_file, send_from_directory, abort, jsonify
from werkzeug.utils import secure_filename

from PIL import Image

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as ExcelImage


app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

UPLOAD_TIME_IN = os.path.join(BASE_DIR, "uploads", "time_in")
UPLOAD_TIME_OUT = os.path.join(BASE_DIR, "uploads", "time_out")
GENERATED_DIR = os.path.join(BASE_DIR, "generated")
LOGO_PATH = os.path.join(BASE_DIR, "static", "henderson_logo.png")
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "attendance.db")

os.makedirs(UPLOAD_TIME_IN, exist_ok=True)
os.makedirs(UPLOAD_TIME_OUT, exist_ok=True)
os.makedirs(GENERATED_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

TIME_IN_RECORDS = []
TIME_OUT_RECORDS = []
JOBS = {}

EMPLOYEES_SEED = [
    ("15829A", "SAMUEL ONG KAI WEN"),
    ("17327", "SANJAY"),
    ("14464", "GENIEVE CAROLE ADRIANO"),
    ("17769", "GOVINDHRAJ MURUGAN"),
]


def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db_connection()

    conn.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_code TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1
        )
    """)

    for emp_code, name in EMPLOYEES_SEED:
        conn.execute("""
            INSERT OR IGNORE INTO employees (emp_code, name, active)
            VALUES (?, ?, 1)
        """, (emp_code, name))

    conn.commit()
    conn.close()


def get_employees():
    conn = get_db_connection()

    employees = conn.execute("""
        SELECT emp_code, name
        FROM employees
        WHERE active = 1
        ORDER BY name ASC
    """).fetchall()

    conn.close()
    return employees


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def save_uploaded_file(file, target_folder):
    original_filename = secure_filename(file.filename)
    ext = original_filename.rsplit(".", 1)[1].lower()
    unique_filename = f"{uuid.uuid4().hex}.{ext}"
    file_path = os.path.join(target_folder, unique_filename)
    file.save(file_path)

    return file_path, original_filename, unique_filename


def extract_datetime_from_filename(filename):
    """
    Extract date/time from WhatsApp-style filename.

    Supported examples before/after secure_filename():
    - WhatsApp Image 2026-05-28 at 18.57.22.jpeg
    - WhatsApp_Image_2026-05-28_at_18.57.22.jpeg
    - WhatsApp Image 2026-05-28 at 18.57.22 (1).jpeg
    - WhatsApp_Image_2026-05-28_at_18.57.22_1.jpeg

    OCR is intentionally not used in this version.
    Date and time are both taken from the filename.
    """
    if not filename:
        return None

    patterns = [
        r"(\d{4})-(\d{2})-(\d{2})[\s_]+at[\s_]+(\d{1,2})[.\-:](\d{2})[.\-:](\d{2})",
        r"(\d{4})-(\d{2})-(\d{2}).*?(\d{1,2})[.\-:](\d{2})[.\-:](\d{2})",
    ]

    for pattern in patterns:
        match = re.search(pattern, filename, flags=re.IGNORECASE)
        if match:
            year, month, day, hour, minute, second = match.groups()

            try:
                return datetime(
                    int(year),
                    int(month),
                    int(day),
                    int(hour),
                    int(minute),
                    int(second),
                )
            except Exception:
                return None

    return None


def format_date(dt):
    if not dt:
        return ""
    return f"{dt.day}/{dt.month}/{dt.year}"


def format_time(dt):
    if not dt:
        return ""
    return dt.strftime("%H:%M:%S")


def build_filename_record(file_path, original_filename, stored_filename, upload_type):
    filename_dt = extract_datetime_from_filename(original_filename)

    if filename_dt:
        display_date = format_date(filename_dt)
        display_time = format_time(filename_dt)
        sort_dt = filename_dt
        source_used = "FILENAME_DATETIME"
        note = "Used filename date and time."
    else:
        display_date = ""
        display_time = ""
        sort_dt = datetime.max
        source_used = "FILENAME_FAILED"
        note = "Filename date/time could not be detected. Manual entry required."

    return {
        "id": uuid.uuid4().hex,
        "original_filename": original_filename,
        "stored_filename": stored_filename,
        "upload_type": upload_type,
        "file_path": file_path,
        "filename_datetime": filename_dt,
        "detected_datetime": sort_dt,
        "date": display_date,
        "time": display_time,
        "ocr_text": note,
        "source_used": source_used,
    }


def process_uploads(files, upload_type):
    """
    Filename-only processing.

    1. Save uploaded files.
    2. Extract date and time from the WhatsApp filename.
    3. Sort chronologically using filename date/time.
    4. No OCR is used.
    """
    records = []

    target_folder = UPLOAD_TIME_IN if upload_type == "time_in" else UPLOAD_TIME_OUT

    for file in files:
        if file and allowed_file(file.filename):
            file_path, original_filename, stored_filename = save_uploaded_file(file, target_folder)

            records.append(
                build_filename_record(
                    file_path=file_path,
                    original_filename=original_filename,
                    stored_filename=stored_filename,
                    upload_type=upload_type,
                )
            )

    records.sort(key=lambda x: x["detected_datetime"] or datetime.max)
    return records


def process_saved_files(saved_files, upload_type, job_id, start_progress, end_progress):
    """
    Filename-only background processing.
    This is very fast because it does not call Tesseract/OCR.
    """
    records = []
    total_files = len(saved_files)

    for index, saved in enumerate(saved_files):
        file_path = saved["file_path"]
        original_filename = saved["original_filename"]
        stored_filename = saved["stored_filename"]

        records.append(
            build_filename_record(
                file_path=file_path,
                original_filename=original_filename,
                stored_filename=stored_filename,
                upload_type=upload_type,
            )
        )

        if total_files > 0:
            progress_range = end_progress - start_progress
            progress = start_progress + int(((index + 1) / total_files) * progress_range)
            JOBS[job_id]["progress"] = progress
            JOBS[job_id]["message"] = (
                f"Reading filename date/time for {upload_type.replace('_', ' ').upper()} photos: "
                f"{index + 1}/{total_files}"
            )

    records.sort(key=lambda x: x["detected_datetime"] or datetime.max)
    return records


def process_job_background(job_id, saved_time_in_files, saved_time_out_files):
    try:
        JOBS[job_id]["message"] = "Reading TIME IN filenames..."
        JOBS[job_id]["progress"] = 5

        time_in_records = process_saved_files(
            saved_time_in_files,
            "time_in",
            job_id,
            5,
            50,
        )

        JOBS[job_id]["message"] = "Reading TIME OUT filenames..."
        JOBS[job_id]["progress"] = 50

        time_out_records = process_saved_files(
            saved_time_out_files,
            "time_out",
            job_id,
            50,
            95,
        )

        JOBS[job_id]["time_in_records"] = time_in_records
        JOBS[job_id]["time_out_records"] = time_out_records
        JOBS[job_id]["progress"] = 100
        JOBS[job_id]["message"] = "Processing completed."
        JOBS[job_id]["status"] = "done"

    except Exception as e:
        JOBS[job_id]["status"] = "error"
        JOBS[job_id]["error"] = str(e)
        JOBS[job_id]["message"] = f"Error: {str(e)}"


def get_sorted_records():
    time_in_sorted = sorted(
        TIME_IN_RECORDS,
        key=lambda x: x["detected_datetime"] or datetime.max,
    )

    time_out_sorted = sorted(
        TIME_OUT_RECORDS,
        key=lambda x: x["detected_datetime"] or datetime.max,
    )

    return time_in_sorted, time_out_sorted


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        site = request.form.get("site", "").strip()
        month_year = request.form.get("month_year", "").strip()

        time_in_files = request.files.getlist("time_in_photos")
        time_out_files = request.files.getlist("time_out_photos")

        job_id = uuid.uuid4().hex

        JOBS[job_id] = {
            "status": "processing",
            "progress": 0,
            "message": "Saving uploaded photos...",
            "site": site,
            "month_year": month_year,
            "time_in_records": [],
            "time_out_records": [],
            "error": None,
        }

        saved_time_in_files = []
        saved_time_out_files = []

        for file in time_in_files:
            if file and allowed_file(file.filename):
                file_path, original_filename, stored_filename = save_uploaded_file(file, UPLOAD_TIME_IN)
                saved_time_in_files.append({
                    "file_path": file_path,
                    "original_filename": original_filename,
                    "stored_filename": stored_filename,
                })

        for file in time_out_files:
            if file and allowed_file(file.filename):
                file_path, original_filename, stored_filename = save_uploaded_file(file, UPLOAD_TIME_OUT)
                saved_time_out_files.append({
                    "file_path": file_path,
                    "original_filename": original_filename,
                    "stored_filename": stored_filename,
                })

        thread = threading.Thread(
            target=process_job_background,
            args=(job_id, saved_time_in_files, saved_time_out_files),
            daemon=True,
        )
        thread.start()

        return redirect(url_for("processing", job_id=job_id))

    return render_template("index.html")


@app.route("/uploaded/<upload_type>/<filename>")
def uploaded_file(upload_type, filename):
    if upload_type == "time_in":
        folder = UPLOAD_TIME_IN
    elif upload_type == "time_out":
        folder = UPLOAD_TIME_OUT
    else:
        abort(404)

    return send_from_directory(folder, filename)


@app.route("/processing/<job_id>")
def processing(job_id):
    if job_id not in JOBS:
        abort(404)

    return render_template("processing.html", job_id=job_id)


@app.route("/job_status/<job_id>")
def job_status(job_id):
    job = JOBS.get(job_id)

    if not job:
        return jsonify({
            "status": "missing",
            "progress": 0,
            "message": "Job not found.",
            "error": "Job not found.",
        }), 404

    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "message": job["message"],
        "error": job["error"],
    })


@app.route("/preview", methods=["GET"])
def preview():
    job_id = request.args.get("job_id")

    if job_id:
        job = JOBS.get(job_id)
        if not job:
            abort(404)

        site = job["site"]
        month_year = job["month_year"]

        time_in_sorted = sorted(
            job["time_in_records"],
            key=lambda x: x["detected_datetime"] or datetime.max,
        )

        time_out_sorted = sorted(
            job["time_out_records"],
            key=lambda x: x["detected_datetime"] or datetime.max,
        )
    else:
        site = request.args.get("site", "")
        month_year = request.args.get("month_year", "")
        time_in_sorted, time_out_sorted = get_sorted_records()

    max_rows = max(len(time_in_sorted), len(time_out_sorted))
    rows = []

    for i in range(max_rows):
        time_in = time_in_sorted[i] if i < len(time_in_sorted) else None
        time_out = time_out_sorted[i] if i < len(time_out_sorted) else None

        date_value = ""
        if time_in and time_in.get("date"):
            date_value = time_in["date"]
        elif time_out and time_out.get("date"):
            date_value = time_out["date"]

        rows.append({
            "index": i,
            "date": date_value,
            "time_in": time_in,
            "time_out": time_out,
            "site": site,
        })

    employees = get_employees()

    return render_template(
        "preview.html",
        rows=rows,
        site=site,
        month_year=month_year,
        employees=employees,
        job_id=job_id,
    )


@app.route("/employees", methods=["GET", "POST"])
def employees_page():
    if request.method == "POST":
        emp_code = request.form.get("emp_code", "").strip()
        name = request.form.get("name", "").strip().upper()

        if emp_code and name:
            conn = get_db_connection()
            conn.execute("""
                INSERT INTO employees (emp_code, name, active)
                VALUES (?, ?, 1)
                ON CONFLICT(emp_code) DO UPDATE SET
                    name = excluded.name,
                    active = 1
            """, (emp_code, name))
            conn.commit()
            conn.close()

        return redirect(url_for("employees_page"))

    search = request.args.get("search", "").strip()
    sort_by = request.args.get("sort_by", "name").strip()
    sort_order = request.args.get("sort_order", "asc").strip().lower()
    per_page = request.args.get("per_page", "10").strip()
    page = request.args.get("page", "1").strip()

    allowed_sort_columns = {
        "emp_code": "emp_code",
        "name": "name",
    }

    if sort_by not in allowed_sort_columns:
        sort_by = "name"

    if sort_order not in ["asc", "desc"]:
        sort_order = "asc"

    try:
        per_page = int(per_page)
    except Exception:
        per_page = 10

    if per_page not in [5, 10, 50]:
        per_page = 10

    try:
        page = int(page)
    except Exception:
        page = 1

    if page < 1:
        page = 1

    conn = get_db_connection()

    where_clause = ""
    params = []

    if search:
        where_clause = "WHERE emp_code LIKE ? OR name LIKE ?"
        search_value = f"%{search}%"
        params.extend([search_value, search_value])

    total_count = conn.execute(
        f"SELECT COUNT(*) FROM employees {where_clause}",
        params,
    ).fetchone()[0]

    total_pages = max((total_count + per_page - 1) // per_page, 1)

    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    order_column = allowed_sort_columns[sort_by]
    order_direction = "ASC" if sort_order == "asc" else "DESC"

    employees = conn.execute(f"""
        SELECT id, emp_code, name
        FROM employees
        {where_clause}
        ORDER BY {order_column} {order_direction}
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()

    conn.close()

    return render_template(
        "employees.html",
        employees=employees,
        search=search,
        sort_by=sort_by,
        sort_order=sort_order,
        per_page=per_page,
        page=page,
        total_pages=total_pages,
        total_count=total_count,
    )


@app.route("/employees/update/<int:employee_id>", methods=["POST"])
def update_employee(employee_id):
    emp_code = request.form.get("emp_code", "").strip()
    name = request.form.get("name", "").strip().upper()

    search = request.form.get("search", "")
    sort_by = request.form.get("sort_by", "name")
    sort_order = request.form.get("sort_order", "asc")
    per_page = request.form.get("per_page", "10")
    page = request.form.get("page", "1")

    if emp_code and name:
        conn = get_db_connection()
        conn.execute("""
            UPDATE employees
            SET emp_code = ?, name = ?
            WHERE id = ?
        """, (emp_code, name, employee_id))
        conn.commit()
        conn.close()

    return redirect(url_for(
        "employees_page",
        search=search,
        sort_by=sort_by,
        sort_order=sort_order,
        per_page=per_page,
        page=page,
    ))


@app.route("/employees/delete/<int:employee_id>", methods=["POST"])
def delete_employee(employee_id):
    search = request.form.get("search", "")
    sort_by = request.form.get("sort_by", "name")
    sort_order = request.form.get("sort_order", "asc")
    per_page = request.form.get("per_page", "10")
    page = request.form.get("page", "1")

    conn = get_db_connection()
    conn.execute("""
        DELETE FROM employees
        WHERE id = ?
    """, (employee_id,))
    conn.commit()
    conn.close()

    return redirect(url_for(
        "employees_page",
        search=search,
        sort_by=sort_by,
        sort_order=sort_order,
        per_page=per_page,
        page=page,
    ))


def resize_image_portrait_for_excel(image_path, max_width=120, max_height=160):
    """
    Keeps the image aspect ratio and fits it inside a portrait-style Excel cell.
    Returns width and height in pixels.
    """
    with Image.open(image_path) as pil_img:
        original_width, original_height = pil_img.size

    if original_width == 0 or original_height == 0:
        return max_width, max_height

    width_ratio = max_width / original_width
    height_ratio = max_height / original_height
    scale = min(width_ratio, height_ratio)

    new_width = int(original_width * scale)
    new_height = int(original_height * scale)

    return new_width, new_height


def resize_logo_for_excel(image_path, max_width=360, max_height=110):
    """
    Keeps logo aspect ratio and fits it into the top merged row.
    """
    with Image.open(image_path) as pil_img:
        original_width, original_height = pil_img.size

    if original_width == 0 or original_height == 0:
        return max_width, max_height

    width_ratio = max_width / original_width
    height_ratio = max_height / original_height
    scale = min(width_ratio, height_ratio)

    new_width = int(original_width * scale)
    new_height = int(original_height * scale)

    return new_width, new_height


@app.route("/generate_excel", methods=["POST"])
def generate_excel():
    month_year = request.form.get("month_year", "").strip()
    site = request.form.get("site", "").strip()
    row_count = int(request.form.get("row_count", 0))

    job_id = request.form.get("job_id", "").strip()

    if job_id and job_id in JOBS:
        job = JOBS[job_id]

        time_in_sorted = sorted(
            job["time_in_records"],
            key=lambda x: x["detected_datetime"] or datetime.max,
        )

        time_out_sorted = sorted(
            job["time_out_records"],
            key=lambda x: x["detected_datetime"] or datetime.max,
        )
    else:
        time_in_sorted, time_out_sorted = get_sorted_records()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    title = f"ATTENDANCE {month_year.upper()}"

    headers = [
        "SHIFT",
        "DATE",
        "PHOTO IN",
        "TIME IN",
        "PHOTO OUT",
        "TIME OUT",
        "EMP CODE",
        "NAME",
        "SITE",
    ]

    # Logo row
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 105

    if os.path.exists(LOGO_PATH):
        logo = ExcelImage(LOGO_PATH)

        logo_width, logo_height = resize_logo_for_excel(
            LOGO_PATH,
            max_width=360,
            max_height=100,
        )

        logo.width = logo_width
        logo.height = logo_height

        ws.add_image(logo, "E1")

    # Title row
    ws.merge_cells("A2:I2")
    ws["A2"] = title
    ws["A2"].font = Font(size=16, bold=True, underline="single", color="000000")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    # Header row
    header_row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="2E75B6")

    widths = {
        "A": 10,
        "B": 15,
        "C": 18,
        "D": 15,
        "E": 18,
        "F": 15,
        "G": 15,
        "H": 32,
        "I": 35,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    thin = Side(border_style="thin", color="B7C9E2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    excel_row = 4

    for i in range(row_count):
        shift = request.form.get(f"shift_{i}", "").strip()
        date_value = request.form.get(f"date_{i}", "").strip()
        time_in_value = request.form.get(f"time_in_{i}", "").strip()
        time_out_value = request.form.get(f"time_out_{i}", "").strip()
        emp_code = request.form.get(f"emp_code_{i}", "").strip()
        name = request.form.get(f"name_{i}", "").strip()
        site_value = request.form.get(f"site_{i}", site).strip()

        ws.cell(row=excel_row, column=1).value = shift
        ws.cell(row=excel_row, column=2).value = date_value
        ws.cell(row=excel_row, column=4).value = time_in_value
        ws.cell(row=excel_row, column=6).value = time_out_value
        ws.cell(row=excel_row, column=7).value = emp_code
        ws.cell(row=excel_row, column=8).value = name
        ws.cell(row=excel_row, column=9).value = site_value

        ws.row_dimensions[excel_row].height = 125

        # Insert Photo In
        if i < len(time_in_sorted):
            img_path = time_in_sorted[i]["file_path"]
            if os.path.exists(img_path):
                img = ExcelImage(img_path)

                new_width, new_height = resize_image_portrait_for_excel(
                    img_path,
                    max_width=115,
                    max_height=160,
                )

                img.width = new_width
                img.height = new_height

                ws.add_image(img, f"C{excel_row}")

        # Insert Photo Out
        if i < len(time_out_sorted):
            img_path = time_out_sorted[i]["file_path"]
            if os.path.exists(img_path):
                img = ExcelImage(img_path)

                new_width, new_height = resize_image_portrait_for_excel(
                    img_path,
                    max_width=115,
                    max_height=160,
                )

                img.width = new_width
                img.height = new_height

                ws.add_image(img, f"E{excel_row}")

        for col in range(1, 10):
            cell = ws.cell(row=excel_row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        excel_row += 1

    # Prepared By row at bottom
    prepared_row = excel_row + 1

    ws.merge_cells(start_row=prepared_row, start_column=1, end_row=prepared_row, end_column=4)

    prepared_cell = ws.cell(row=prepared_row, column=1)
    prepared_cell.value = "Prepared By : Thaqif Kasmani"
    prepared_cell.font = Font(size=12, bold=False, color="000000")
    prepared_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[prepared_row].height = 25

    for col in range(1, 5):
        cell = ws.cell(row=prepared_row, column=col)
        cell.border = border

    output_filename = f"attendance_{uuid.uuid4().hex}.xlsx"
    output_path = os.path.join(GENERATED_DIR, output_filename)

    wb.save(output_path)

    return send_file(output_path, as_attachment=True, download_name="attendance.xlsx")


if __name__ == "__main__":
    init_db()
    app.run(debug=True)
